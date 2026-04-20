#Author = Shalabodov Yaroslav Dmitrievich
#Group = P3110
#Date =  19.11.25

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

df = pd.read_excel("lab5.xlsx",usecols=[i for i in range(25)], skiprows=2, nrows = 12)

df = df.drop(columns=df.columns[5])

df.to_excel("Output.xlsx", index=False, header=False)

wb = load_workbook("Output.xlsx")
ws = wb.active

thin_border = Border(
    left=Side(style='thick', color='FF0000'),
    right=Side(style='thick',color='FF0000'),
    top=Side(style='thick',color='FF0000'),
    bottom=Side(style='thick',color='FF0000')
)

top = Border(
    top=Side(style='thick', color='FF0000')
)

bottom = Border(
    bottom=Side(style='thick',color='FF0000')
)

right = Border(
    right=Side(style='thick', color='FF0000')
)

left = Border(
    left=Side(style='thick', color='FF0000')
)

leftUpCorner = Border(
    left=Side(style='thick', color='FF0000'),
    top=Side(style='thick', color='FF0000')
)
leftDownCorner = Border(
    left=Side(style='thick', color='FF0000'),
    bottom=Side(style='thick', color='FF0000')
)
rightUpCorner = Border(
    right=Side(style='thick', color='FF0000'),
    top=Side(style='thick', color='FF0000')
)
rightDownCorner = Border(
    right=Side(style='thick', color='FF0000'),
    bottom=Side(style='thick', color='FF0000')
)

for row in ws.iter_rows():
    for cell in row:
        if cell.column in [5,24]:
            cell.border = right
        elif cell.column == 4:
            cell.border = left
        elif cell.row == 1 and cell.column in [i for i in range(6,25)]:
            cell.border = top
        elif cell.row == 12 and cell.column in [i for i in range(6,25)]:
            cell.border = bottom
        elif not (cell.column in [i for i in range(6,25)] and cell.row in [i for i in range(2,12)]):
            cell.border = thin_border

#Добавляем углы
ws[12][3].border = leftDownCorner
ws[1][3].border = leftUpCorner
ws[1][4].border = rightUpCorner
ws[1][23].border = rightUpCorner
ws[12][4].border =rightDownCorner
ws[12][23].border = rightDownCorner


for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        except:
            pass

    adjusted_width = min(max_length + 2, 52)
    ws.column_dimensions[column_letter].width = adjusted_width

wb.save("Output.xlsx")